from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from dotenv import load_dotenv
import markdown
import os
import io
import re

load_dotenv()

app = FastAPI(title="Excel Markdown to HTML Converter")

# Get origins from environment variable, default to ["*"] if not set
origins_str = os.getenv("ALLOWED_ORIGINS", "*")
if origins_str == "*":
    origins = ["*"]
else:
    origins = [origin.strip() for origin in origins_str.split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://kcglobed-lms-admin.web.app",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def text_to_html(text):
    if text is None:
        return None

    print(f"\n--- PROCESSING NEW CELL ---")
    print(f"RAW EXCEL TEXT (repr): {repr(text)}")

    text_str = str(text)
    # Excel sometimes uses \r (carriage return) or \x0b (vertical tab) instead of \n for soft line breaks.
    # We must normalize all these weird characters into standard \n first.
    text_str = text_str.replace('\r\n', '\n')
    text_str = text_str.replace('\r', '\n')
    text_str = text_str.replace('\x0b', '\n')
    text_str = text_str.replace('\u2028', '\n')

    # Now that we only have \n, we replace 1 or more newlines with double newlines
    # so that Markdown wraps each line block in its own <p> tag.
    text_str = re.sub(r'\n+', '\n\n', text_str).strip()

    print(f"NORMALIZED TEXT (repr): {repr(text_str)}")

    html_output = markdown.markdown(
        text_str,
        extensions=[
            "tables",
            "fenced_code",
        ],
    )

    print(f"FINAL HTML OUTPUT: {repr(html_output)}")
    print(f"---------------------------\n")

    return html_output


@app.post("/convert")
async def convert_excel(file: UploadFile = File(...)):
    try:

        if not file.filename.endswith(".xlsx"):
            raise HTTPException(
                status_code=400,
                detail="Please upload only .xlsx file"
            )

        # Read uploaded file
        contents = await file.read()

        # Load workbook
        wb = load_workbook(io.BytesIO(contents))

        # Convert specific columns
        target_columns = {"question", "solution"}

        for sheet in wb.worksheets:
            col_indices = []

            for i, row in enumerate(sheet.iter_rows()):
                if i == 4:
                    # Parse header row (Row 5 is index 4)
                    for idx, cell in enumerate(row):
                        if isinstance(cell.value, str) and cell.value.strip().lower() in target_columns:
                            col_indices.append(idx)
                    continue
                
                # Process data rows for only the target columns
                # (Rows 1-4 will naturally be skipped because col_indices is empty until row 5)
                for idx in col_indices:
                    if idx < len(row):
                        cell = row[idx]
                        if isinstance(cell.value, str):
                            cell.value = text_to_html(cell.value)

        # Save the modified workbook to a new in-memory byte stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Output filename
        output_filename = f"converted_{file.filename}"

        headers = {
            'Content-Disposition': f'attachment; filename="{output_filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }

        # Return downloadable file from memory
        from fastapi.responses import Response
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )